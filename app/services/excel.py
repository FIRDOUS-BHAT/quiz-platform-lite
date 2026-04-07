from io import BytesIO
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - environment-dependent import guard
    load_workbook = None

from app.schemas.quiz import QuizDefinition
from app.utils.time import coerce_epoch


def slugify(value: str) -> str:
    normalized = "".join(char.lower() if char.isalnum() else "-" for char in value.strip())
    compact = "-".join(part for part in normalized.split("-") if part)
    return compact[:64] or "quiz"


def parse_quiz_workbook(content: bytes, filename: str | None = None) -> QuizDefinition:
    if load_workbook is None:
        raise RuntimeError("Excel import requires the optional dependency 'openpyxl'")
    workbook = load_workbook(BytesIO(content), data_only=True)
    metadata_sheet = _find_sheet(workbook, "metadata")
    questions_sheet = _find_sheet(workbook, "questions")

    if questions_sheet is None:
        raise ValueError("Workbook must contain a 'questions' sheet")

    metadata = _read_metadata(metadata_sheet, filename)
    questions = _read_questions(questions_sheet)

    if not questions:
        raise ValueError("Workbook must include at least one question row")

    quiz_id = metadata.get("quiz_id") or slugify(metadata["title"])

    return QuizDefinition.model_validate(
        {
            "quiz_id": quiz_id,
            "version": metadata.get("version") or "1",
            "title": metadata["title"],
            "description": metadata.get("description"),
            "duration_seconds": metadata["duration_seconds"],
            "availability_start_at": metadata.get("availability_start_at"),
            "availability_end_at": metadata.get("availability_end_at"),
            "questions": questions,
        }
    )


def _find_sheet(workbook, target: str):
    target_lower = target.lower()
    for name in workbook.sheetnames:
        if name.lower() == target_lower:
            return workbook[name]
    return None


def _read_metadata(sheet, filename: str | None) -> dict[str, object]:
    default_title = Path(filename or "quiz.xlsx").stem.replace("_", " ").strip() or "Quiz"
    metadata: dict[str, object] = {
        "title": default_title,
        "duration_seconds": 1800,
    }

    if sheet is None:
        return metadata

    for raw_key, raw_value, *_ in sheet.iter_rows(min_row=1, values_only=True):
        if raw_key is None or raw_value is None:
            continue
        key = str(raw_key).strip().lower()
        value = raw_value.strip() if isinstance(raw_value, str) else raw_value
        metadata[key] = value

    if "duration_seconds" in metadata:
        metadata["duration_seconds"] = int(metadata["duration_seconds"])
    for field_name in ("availability_start_at", "availability_end_at"):
        if field_name in metadata:
            metadata[field_name] = coerce_epoch(metadata[field_name], field_name=field_name)

    return metadata


def _read_questions(sheet) -> list[dict[str, object]]:
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("Questions sheet is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    if "correct_option" not in headers:
        raise ValueError("Questions sheet must include a 'correct_option' column")

    option_columns = [
        (index, header.replace("option_", ""))
        for index, header in enumerate(headers)
        if header.startswith("option_")
    ]
    if len(option_columns) < 2:
        raise ValueError("Questions sheet must define at least two option columns")

    question_id_index = headers.index("question_id") if "question_id" in headers else None
    prompt_index = None
    for candidate in ("question_text", "prompt", "text"):
        if candidate in headers:
            prompt_index = headers.index(candidate)
            break
    if prompt_index is None:
        raise ValueError("Questions sheet must include 'question_text', 'prompt', or 'text'")

    correct_index = headers.index("correct_option")

    questions: list[dict[str, object]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        prompt = row[prompt_index]
        if prompt in (None, ""):
            continue

        options = []
        for column_index, option_id in option_columns:
            value = row[column_index] if column_index < len(row) else None
            if value not in (None, ""):
                options.append({"id": option_id, "text": str(value).strip()})

        if len(options) < 2:
            raise ValueError(f"Row {row_number} must include at least two options")

        correct_option = row[correct_index]
        if correct_option in (None, ""):
            raise ValueError(f"Row {row_number} is missing correct_option")

        correct_option_normalized = str(correct_option).strip().lower()
        option_ids = {opt["id"] for opt in options}
        if correct_option_normalized not in option_ids:
            raise ValueError(
                f"Row {row_number}: correct_option '{correct_option_normalized}' "
                f"does not match any option column ({', '.join(sorted(option_ids))})"
            )

        question_id = row[question_id_index] if question_id_index is not None else None
        if question_id in (None, ""):
            question_id = f"q{len(questions) + 1}"

        questions.append(
            {
                "id": str(question_id).strip(),
                "prompt": str(prompt).strip(),
                "type": "single_choice",
                "options": options,
                "correct_option_id": correct_option_normalized,
            }
        )

    return questions
